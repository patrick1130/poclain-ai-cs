from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form, Query
from sqlalchemy.orm import Session, defer
from sqlalchemy import func, desc
from typing import List, Optional
from pydantic import BaseModel
from datetime import datetime
import asyncio
from concurrent.futures import ProcessPoolExecutor
import logging
import io
import re

from ..core.database import get_db
from ..models.database import KnowledgeDoc
from ..core.config import settings
from ..utils.vector_db import VectorDB
from ..utils.document_processor import split_document

logger = logging.getLogger(__name__)
router = APIRouter()

process_pool = ProcessPoolExecutor(max_workers=2)


class KnowledgeDocResponse(BaseModel):
    id: int
    title: str
    category: str
    version: int
    create_time: datetime
    update_time: Optional[datetime] = None

    class Config:
        from_attributes = True


class KnowledgeDocDetailResponse(KnowledgeDocResponse):
    content: str


def _get_vector_db():
    return VectorDB(settings.VECTOR_DB_PATH)


# ==========================================
# 🚨 静态路由：必须排在动态路由 /{doc_id} 之前
# ==========================================


@router.get("/docs", response_model=List[KnowledgeDocResponse])
def get_knowledge_docs(
    category: Optional[str] = Query(None),
    keyword: Optional[str] = Query(None),
    db: Session = Depends(get_db),
):
    # 🚨 架构师加固：强制执行 defer('content')，防止百兆级 Markdown 并发撑爆内存
    query = db.query(KnowledgeDoc).options(defer(KnowledgeDoc.content))
    if category:
        query = query.filter(KnowledgeDoc.category == category)
    if keyword:
        query = query.filter(KnowledgeDoc.title.contains(keyword))
    return query.order_by(desc(KnowledgeDoc.create_time)).all()


@router.get("/categories")
def get_categories(db: Session = Depends(get_db)):
    categories = db.query(KnowledgeDoc.category).distinct().all()
    return [c[0] for c in categories if c and c[0]]


# ==========================================
# 🚨 核心加固：高性能流式 Excel 解析引擎
# ==========================================


def _extract_text_from_excel(excel_bytes: bytes) -> str:
    try:
        import pandas as pd
        from openpyxl import load_workbook

        excel_file = io.BytesIO(excel_bytes)
        wb = load_workbook(excel_file, read_only=True, keep_links=False)
        sheet_names = wb.sheetnames
        wb.close()

        text_content = []

        with pd.ExcelFile(excel_file, engine="openpyxl") as xls:
            for sheet_name in sheet_names:
                logger.info(f"📊 正在解析 Sheet: {sheet_name}")
                df = pd.read_excel(xls, sheet_name=sheet_name)
                if df.empty:
                    continue

                text_content.append(f"### 【工作表: {sheet_name}】")
                df.dropna(how="all", inplace=True)
                text_content.append(df.to_markdown(index=False))

        excel_file.close()
        return "\n\n".join(text_content)
    except Exception as e:
        logger.error(f"❌ Excel 解析崩溃: {str(e)}")
        raise ValueError(f"Excel 解析失败: {str(e)}")


def _extract_text_from_pdf(pdf_bytes: bytes) -> str:
    try:
        import fitz

        text_content = []
        with fitz.open(stream=pdf_bytes, filetype="pdf") as doc:
            for page in doc:
                text_content.append(page.get_text())
        return "\n".join(text_content)
    except Exception as e:
        raise ValueError(f"PDF 解析失败: {e}")


# ==========================================
# 🚨 业务逻辑路由
# ==========================================


@router.post("/upload")
async def upload_document(
    file: UploadFile = File(...),
    category: str = Form(...),
    db: Session = Depends(get_db),
):
    file_ext = file.filename.rsplit(".", 1)[-1].lower()
    content_bytes = await file.read()
    title = file.filename.rsplit(".", 1)[0]
    loop = asyncio.get_running_loop()
    v_db = _get_vector_db()

    try:
        if file_ext == "pdf":
            content = await loop.run_in_executor(
                process_pool, _extract_text_from_pdf, content_bytes
            )
        elif file_ext in ["xlsx", "xls"]:
            content = await loop.run_in_executor(
                process_pool, _extract_text_from_excel, content_bytes
            )
        else:
            content = content_bytes.decode("utf-8", errors="ignore")
    except Exception as e:
        logger.error(f"🔥 文件解析失败: {str(e)}")
        raise HTTPException(status_code=400, detail=str(e))

    existing = (
        db.query(KnowledgeDoc)
        .filter(KnowledgeDoc.title == title)
        .order_by(desc(KnowledgeDoc.version))
        .first()
    )
    new_version = (existing.version + 1) if existing else 1

    # 🚨 架构师修复：先落库，锁定唯一物理主键 ID
    db_doc = KnowledgeDoc(
        title=title, content=content, category=category, version=new_version
    )
    db.add(db_doc)
    db.commit()
    db.refresh(db_doc)

    try:
        # 🚨 强一致性防御：严格绑定 db_doc.id 作为向量引擎的生命周期锁
        if existing:
            await v_db.delete_document(existing.id)

        chunks = await loop.run_in_executor(
            process_pool, split_document, title, content, category
        )

        # 将 title 纠正为 db_doc.id，打通双库一致性
        await v_db.add_document(
            doc_id=db_doc.id, title=title, chunks=chunks, category=category
        )
    except Exception as e:
        logger.error(f"⚠️ 向量注入失败，触发 SAGA 补偿事务: {e}")
        # 🚨 灾难回滚：一旦向量库报错，必须删除刚插入的 MySQL 数据，绝不能留下空壳
        db.delete(db_doc)
        db.commit()
        raise HTTPException(
            status_code=500, detail=f"向量知识库构建失败，事务已回滚: {str(e)}"
        )

    return {
        "id": db_doc.id,
        "title": title,
        "category": category,
        "version": new_version,
    }


@router.get("/docs/{doc_id}", response_model=KnowledgeDocDetailResponse)
def get_knowledge_doc_detail(doc_id: int, db: Session = Depends(get_db)):
    db_doc = db.query(KnowledgeDoc).filter(KnowledgeDoc.id == doc_id).first()
    if not db_doc:
        raise HTTPException(status_code=404, detail="文档未找到")
    return db_doc


# 🚨 架构师核心修复：对齐前端资源路由，将 /{doc_id} 修改为 /docs/{doc_id}
@router.delete("/docs/{doc_id}")
async def delete_knowledge_doc(doc_id: int, db: Session = Depends(get_db)):
    db_doc = db.query(KnowledgeDoc).filter(KnowledgeDoc.id == doc_id).first()
    if not db_doc:
        raise HTTPException(status_code=404, detail="未找到文档")
    v_db = _get_vector_db()

    try:
        # 🚨 安全合规审计：必须先彻底物理粉碎底层向量引擎中的绝密数据片段
        await v_db.delete_document(db_doc.id)

        # 随后才允许从关系型元数据表中抹除
        db.delete(db_doc)
        db.commit()
        return {"success": True}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"分布式删除事务崩溃: {str(e)}")
